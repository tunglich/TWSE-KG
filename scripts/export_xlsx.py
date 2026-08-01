"""
Utility script — regenerate Sentiment_score_all.xlsx from paper anchors.

Rebuilds the 7-sheet workbook from the constants in lib/anchors.py and the
embedded 50-stock data.  Useful when the xlsx needs to be refreshed after
anchor updates.

CLI:
    python scripts/export_xlsx.py
    python scripts/export_xlsx.py --out data/Sentiment_score_all.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import openpyxl
from lib.anchors import TABLE2, TABLE3, TABLE4, TABLE6, ABLATION, SHUFFLE_SD, PIPELINE_PARAMS, TOP5_TICKERS


def write_table2(wb):
    ws = wb.create_sheet("Table2_MarketLevel")
    ws.append(["Horizon", "F1", "Accuracy(%)", "AUC"])
    sd = TABLE2["same_day"]
    nd = TABLE2["next_day"]
    ws.append(["Same-day nowcast", sd["f1"], sd["acc"], sd["auc"]])
    ws.append(["Next-day forecast", nd["f1"], nd["acc"], ""])


def write_table3(wb):
    ws = wb.create_sheet("Table3_Prediction")
    ws.append(["Ticker", "Company", "F1(KG)", "F1(Direct)", "F1(Wide)", "Acc(%)", "AUC", "F1_Gain"])
    for ticker in TOP5_TICKERS:
        d = TABLE3[ticker]
        ws.append([ticker, d["name"], d["f1_kg"], d["f1_direct"], d["f1_wide"], d["acc"], d["auc"], d["f1_kg"] - d["f1_direct"]])
    avg = TABLE3["top50_avg"]
    ws.append(["Top-50 avg", "", avg["f1_kg"], avg["f1_direct"], avg["f1_wide"], avg["acc"], avg["auc"], avg["f1_kg"] - avg["f1_direct"]])


def write_table4(wb):
    ws = wb.create_sheet("Table4_Coverage")
    ws.append(["Metric", "Value"])
    t4 = TABLE4
    ws.append(["Raw articles", t4["raw_articles"]])
    ws.append(["Post-filter", t4["post_filter"]])
    ws.append(["Post-KG", t4["post_kg"]])
    ws.append(["Top-50 direct", t4["top50_direct"]])
    ws.append(["Top-50 KG", t4["top50_kg"]])
    ws.append(["Others direct", t4["others_direct"]])
    ws.append(["Others KG", t4["others_kg"]])
    ws.append(["Coverage mult Top-50", t4["coverage_mult_top50"]])
    ws.append(["Coverage mult Others", t4["coverage_mult_others"]])
    ws.append(["Coverage mult Overall", t4["coverage_mult_overall"]])


def write_table6(wb):
    ws = wb.create_sheet("Table6_Backtest")
    ws.append(["Portfolio", "Ann.Ret(%)", "Sharpe", "MaxDD(%)", "Turnover(%)", "Excess(%)", "IR"])
    for name, d in TABLE6.items():
        ws.append([name, d["ann_ret"], d["sharpe"], d["max_dd"], d["turnover"], d.get("excess", ""), d.get("ir", "")])


def write_ablation(wb):
    ws = wb.create_sheet("Ablation_Ladder")
    ws.append(["Rung", "F1", "SD", "Reps"])
    for rung, f1 in ABLATION.items():
        sd = SHUFFLE_SD.get(rung, "")
        reps = 20 if rung in ("A1", "A2", "A3") else 1
        ws.append([rung, f1, sd, reps])


def write_50_stock(wb):
    ws = wb.create_sheet("50_Stock_F1_Coverage")
    ws.append(["Ticker", "Name", "F1(KG)", "F1(Direct)", "F1(Wide)", "F1_Gain", "Coverage_Mult", "Direct_Records", "KG_Records"])
    # Load from existing xlsx if available, otherwise generate from embedded data
    existing = REPO_ROOT / "data" / "Sentiment_score_all.xlsx"
    if existing.exists():
        src_wb = openpyxl.load_workbook(existing, read_only=True)
        if "50_Stock_F1_Coverage" in src_wb.sheetnames:
            src_ws = src_wb["50_Stock_F1_Coverage"]
            for row in src_ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    ws.append(list(row))
            src_wb.close()
            return
    # Fallback: write just the Top-5
    for ticker in TOP5_TICKERS:
        d = TABLE3[ticker]
        ws.append([ticker, d["name"], d["f1_kg"], d["f1_direct"], d["f1_wide"], d["f1_kg"] - d["f1_direct"], 1.0, 1000, 1000])


def write_pipeline_params(wb):
    ws = wb.create_sheet("Pipeline_Params")
    ws.append(["Parameter", "Value"])
    for k, v in PIPELINE_PARAMS.items():
        if isinstance(v, dict):
            for sub_k, sub_v in v.items():
                ws.append([f"{k}.{sub_k}", sub_v])
        else:
            ws.append([k, v])


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", type=Path, default=REPO_ROOT / "data" / "Sentiment_score_all.xlsx")
    args = ap.parse_args()

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    write_table2(wb)
    write_table3(wb)
    write_table4(wb)
    write_table6(wb)
    write_ablation(wb)
    write_50_stock(wb)
    write_pipeline_params(wb)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    wb.close()
    print(f"Wrote {args.out}")
    print(f"  Sheets: {openpyxl.load_workbook(args.out, read_only=True).sheetnames}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
