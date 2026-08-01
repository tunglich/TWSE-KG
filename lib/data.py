"""
Data loading helpers for TWSE-KG pipeline.

Reads Sentiment_score_all.xlsx and exposes typed accessors for each sheet.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = REPO_ROOT / "data" / "Sentiment_score_all.xlsx"

EXPECTED_SHEETS = [
    "Table3_Prediction",
    "Table4_Coverage",
    "Table6_Backtest",
    "Ablation_Ladder",
    "50_Stock_F1_Coverage",
    "Table2_MarketLevel",
    "Pipeline_Params",
]


def load_workbook(path: Path | None = None) -> openpyxl.Workbook:
    """Open the experiment data workbook (read-only)."""
    p = path or DEFAULT_XLSX
    if not p.exists():
        raise FileNotFoundError(f"Data file not found: {p}")
    wb = openpyxl.load_workbook(p, read_only=True)
    missing = set(EXPECTED_SHEETS) - set(wb.sheetnames)
    if missing:
        raise ValueError(f"Missing sheets in {p}: {missing}")
    return wb


def read_sheet_rows(path: Path | None = None, sheet: str = "") -> list[dict[str, Any]]:
    """Read a sheet as a list of dicts (header row → keys)."""
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet}' not in workbook; available: {wb.sheetnames}")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        result.append(dict(zip(headers, row)))
    wb.close()
    return result


def load_table3(path: Path | None = None) -> list[dict[str, Any]]:
    return read_sheet_rows(path, "Table3_Prediction")


def load_table4(path: Path | None = None) -> list[dict[str, Any]]:
    return read_sheet_rows(path, "Table4_Coverage")


def load_table6(path: Path | None = None) -> list[dict[str, Any]]:
    return read_sheet_rows(path, "Table6_Backtest")


def load_ablation(path: Path | None = None) -> list[dict[str, Any]]:
    return read_sheet_rows(path, "Ablation_Ladder")


def load_50_stock(path: Path | None = None) -> list[dict[str, Any]]:
    rows = read_sheet_rows(path, "50_Stock_F1_Coverage")
    # Filter out the summary "Average" row at the bottom
    return [r for r in rows if str(r.get("Ticker", "")).strip() not in ("Average", "", "None")]


def load_table2(path: Path | None = None) -> list[dict[str, Any]]:
    return read_sheet_rows(path, "Table2_MarketLevel")


def load_pipeline_params(path: Path | None = None) -> list[dict[str, Any]]:
    return read_sheet_rows(path, "Pipeline_Params")
